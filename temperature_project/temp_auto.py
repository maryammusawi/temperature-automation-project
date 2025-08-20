import requests, time, os
from openpyxl import load_workbook

def update_temps():
    wb = load_workbook('cities.xlsx')
    sheet = wb.active
    message = "🌡️ Temperature Update:\n\n"
    
    for row in range(2, sheet.max_row + 1):
        city = sheet[f'A{row}'].value
        if city:
            try:
                weather = requests.get(f"http://api.openweathermap.org/data/2.5/weather?q={city}&units=metric&appid=e969633cc002095a8bc66089dec4a71b").json()
                temp = weather['main']['temp']
                description = weather['weather'][0]['description'].title()
                
                sheet[f'B{row}'] = temp
                sheet[f'C{row}'] = description
                
                if temp > 30:
                    os.system(f"""osascript -e 'display notification "🔥 Hot in {city}! {temp}°C"'""")
                elif temp < 5:
                    os.system(f"""osascript -e 'display notification "❄️ Cold in {city}! {temp}°C"'""")
                
                message += f"• {city}: {temp}°C, {description}\n"
                
            except Exception as e:
                sheet[f'B{row}'] = "Error"
                sheet[f'C{row}'] = "Error"
                message += f"• {city}: Error\n"
    
    wb.save('updated_temps.xlsx')
    os.system(f"""osascript -e 'display notification "Temperatures updated!"'""")
    print(f"✅ Updated at {time.strftime('%H:%M')}")

print("🌤️ Temperature tracker started (8 AM daily)")
while True:
    if time.strftime("%H:%M") == "08:00":
        update_temps()
    time.sleep(60)